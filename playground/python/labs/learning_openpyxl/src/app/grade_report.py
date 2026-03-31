"""
grade_report.py

このモジュールでは、Python 側で計算した結果を Excel に書き出す題材を扱います。
単なるセル書き込みではなく、合計や平均、判定のような「意味のある列」を作ることで、
openpyxl を業務データの整形に使うイメージを掴みやすくしています。
"""

from collections.abc import Sequence
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


@dataclass
class StudentScore:
    """学習用の簡易成績データです。"""

    name: str
    japanese: int
    math: int
    science: int


def build_grade_report(scores: Sequence[StudentScore]) -> Workbook:
    """
    成績表の Workbook を作成して返します。

    openpyxl は Excel 計算エンジンではないため、
    学習の初期段階では Python 側で値を計算してからセルへ書き込む方が理解しやすいです。
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Grades"

    headers = ("氏名", "国語", "数学", "理科", "合計", "平均", "判定")
    sheet.append(headers)

    review_fill = PatternFill(fill_type="solid", fgColor="FDE2E4")
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    for score in scores:
        total = score.japanese + score.math + score.science
        average = round(total / 3, 1)
        result = "合格" if average >= 70 else "要復習"

        sheet.append(
            [
                score.name,
                score.japanese,
                score.math,
                score.science,
                total,
                average,
                result,
            ]
        )

        current_row = sheet.max_row
        sheet[f"F{current_row}"].number_format = "0.0"
        if result == "要復習":
            sheet[f"G{current_row}"].fill = review_fill

    sheet.freeze_panes = "A2"
    return workbook


def save_grade_report(path: Path, scores: Sequence[StudentScore]) -> None:
    """
    成績表を xlsx ファイルとして保存します。

    保存先を関数引数にしておくことで、
    テストでは tmp_path を使って安全に検証できます。
    """
    workbook = build_grade_report(scores)
    workbook.save(path)