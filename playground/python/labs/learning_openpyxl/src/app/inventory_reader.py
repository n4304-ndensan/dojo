"""
inventory_reader.py

このモジュールでは、保存済みの xlsx ファイルを読み戻し、
業務ロジックに使える Python データへ変換する流れを扱います。
openpyxl は書き込みだけでなく、既存ファイルの読み取りにもよく使われます。
"""

from collections.abc import Sequence
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook, load_workbook


@dataclass
class InventoryItem:
    """在庫表の1行を表す簡易データです。"""

    item_name: str
    stock: int
    reorder_level: int


def save_inventory_workbook(path: Path, items: Sequence[InventoryItem]) -> None:
    """在庫データを持つ Workbook を作成し、ファイルへ保存します。"""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Inventory"

    sheet.append(["商品名", "在庫数", "発注点"])
    for item in items:
        sheet.append([item.item_name, item.stock, item.reorder_level])

    workbook.save(path)


def find_low_stock_items(path: Path) -> list[InventoryItem]:
    """
    発注点を下回る商品だけを読み取って返します。

    Excel の行データをそのまま扱うのではなく、
    Python の dataclass に戻すことで後続の処理を書きやすくしています。
    """
    workbook = load_workbook(path)
    sheet = workbook["Inventory"] if "Inventory" in workbook.sheetnames else workbook.active

    low_stock_items: list[InventoryItem] = []
    for item_name, stock, reorder_level in sheet.iter_rows(min_row=2, values_only=True):
        if item_name is None or stock is None or reorder_level is None:
            continue

        stock_value = int(stock)
        reorder_level_value = int(reorder_level)
        if stock_value < reorder_level_value:
            low_stock_items.append(
                InventoryItem(
                    item_name=str(item_name),
                    stock=stock_value,
                    reorder_level=reorder_level_value,
                )
            )

    return low_stock_items