"""
workbook_builder.py

このモジュールでは、openpyxl の最初の題材として、
メモリ上で Workbook を組み立てる処理を扱います。
ファイル保存の前に、「シートを作る」「セルへ値を書く」「簡単な書式を付ける」
という基本操作を確認することが目的です。
"""

from collections.abc import Sequence

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


SalesRow = dict[str, int | str]


def build_sales_workbook(rows: Sequence[SalesRow]) -> Workbook:
    """
    売上一覧の Workbook を作成して返します。

    行データから売上金額を計算し、Workbook に書き込みます。
    返り値をそのままテストできるため、まずは save せずに扱う題材にしています。
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sales"

    headers = ("商品名", "数量", "単価", "売上")
    sheet.append(headers)

    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    for row in rows:
        quantity = int(row["quantity"])
        unit_price = int(row["unit_price"])
        amount = quantity * unit_price
        sheet.append([row["item_name"], quantity, unit_price, amount])

    sheet.freeze_panes = "A2"
    sheet.column_dimensions["A"].width = 18
    sheet.column_dimensions["B"].width = 10
    sheet.column_dimensions["C"].width = 10
    sheet.column_dimensions["D"].width = 12

    for row in sheet.iter_rows(min_row=2, min_col=3, max_col=4):
        for cell in row:
            cell.number_format = "#,##0"

    return workbook