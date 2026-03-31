"""
test_inventory_reader.py

このファイルでは、保存済みの Excel ファイルを load_workbook で読み込み、
必要な行だけを業務データへ変換する流れを学びます。
"""

from pathlib import Path

from openpyxl import load_workbook

from src.app.inventory_reader import InventoryItem, find_low_stock_items, save_inventory_workbook


def test_save_inventory_workbook_creates_expected_sheet(
    tmp_path: Path,
    inventory_items: list[InventoryItem],
) -> None:
    """保存した在庫表のシート名とヘッダーを確認します。"""
    path = tmp_path / "inventory.xlsx"

    save_inventory_workbook(path, inventory_items)
    workbook = load_workbook(path)
    sheet = workbook["Inventory"]

    assert sheet["A1"].value == "商品名"
    assert sheet["B1"].value == "在庫数"
    assert sheet["C1"].value == "発注点"


def test_find_low_stock_items_returns_only_shortage_rows(
    inventory_file: Path,
) -> None:
    """発注点を下回る商品だけが返ることを確認します。"""
    low_stock_items = find_low_stock_items(inventory_file)

    assert low_stock_items == [
        InventoryItem(item_name="Notebook", stock=3, reorder_level=5),
        InventoryItem(item_name="Marker", stock=1, reorder_level=4),
    ]


def test_find_low_stock_items_returns_empty_when_all_items_are_safe(
    tmp_path: Path,
) -> None:
    """すべての在庫が十分な場合は空配列になることを確認します。"""
    path = tmp_path / "safe_inventory.xlsx"
    save_inventory_workbook(
        path,
        [
            InventoryItem(item_name="Tape", stock=8, reorder_level=3),
            InventoryItem(item_name="Clip", stock=6, reorder_level=2),
        ],
    )

    assert find_low_stock_items(path) == []